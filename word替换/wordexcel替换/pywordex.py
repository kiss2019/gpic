# https://blog.csdn.net/qq_40235133/article/details/108762855

from docxtpl import DocxTemplate, RichText
from openpyxl import load_workbook
import os

# 创建要保存的文件夹
def savePath():
    # path="jpg" #jpg为保存目录，可随意更改
    path= r"C:\pycode\wordmodel"
    #如果不存在，新建一个
    if not os.path.exists(path):
          os.makedirs(path)
    os.chdir(path) #进入当前目录
    # print(os.getcwd()) #得到当前目录

def replace(obj):
    if obj is None:
        obj = ''
        return obj


# 加载要填入的数据
wb = load_workbook(r"C:\输入数据.xlsx")
ws = wb['Sheet1']
contexts = []
for row in range(2, ws.max_row + 1):
    name = ws["A" + str(row)].value
    c_name = ws["B" + str(row)].value
    code = ws["C" + str(row)].value
    num = ws["D" + str(row)].value
    time = ws["E" + str(row)].value
    time = str(time)[:-9]
    money = ws["F" + str(row)].value
    address = ws["G" + str(row)].value
    replace_peo = ws["H" + str(row)].value

    context = {"name": name, "c_name": c_name, "code": code, "num": num, "time": time,
               "money": money, "address": address, "replace_peo": replace_peo}
    contexts.append(context)
contexts

savePath()

for context in contexts:
    print(context)
    tpl = DocxTemplate(r"C:\打印模板.docx")
    tpl.render(context)
    tpl.save("./{}的劳动合同.docx".format(context["name"]))